
import io
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import pdfplumber
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

st.set_page_config(page_title="Prop 65 Rule-Based IMDS Analyzer", layout="wide")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CAS_RE = re.compile(r"\b\d{2,7}-\d{2}-\d\b")


def normalize_text(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()


def extract_pages(file_bytes: bytes) -> List[str]:
    pages: List[str] = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            pages.append(txt)
    return pages


def extract_tables(file_bytes: bytes) -> List[List[List[str]]]:
    all_tables = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for t in tables:
                all_tables.append(t)
    return all_tables


def find_metadata(pages: List[str]) -> Dict[str, str]:
    blob = "\n".join(pages[:3])

    def grab(pattern: str) -> str:
        m = re.search(pattern, blob, re.IGNORECASE)
        return normalize_text(m.group(1)) if m else ""

    return {
        "document": "Uploaded MDS Report PDF",
        "part_no": grab(r"Part/Item No\.\:\s*([^\n]+)"),
        "description": grab(r"Description\:\s*([^\n]+)"),
        "supplier": grab(r"Name \[ID\]\:\s*([^\n]+)"),
    }


def load_rules(xlsx_bytes: bytes) -> Tuple[pd.DataFrame, pd.DataFrame]:
    xls = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    rules = pd.read_excel(xls, "rules")
    refs = pd.read_excel(xls, "references")
    rules["CAS No."] = rules["CAS No."].astype(str).str.strip()
    return rules, refs


def infer_context(component: str, material: str, line: str) -> str:
    text = f"{component} {material} {line}".lower()
    if any(k in text for k in ["booster", "propellant", "pyrotechnic"]):
        return "booster_propellant"
    if any(k in text for k in ["label", "ink", "acrylic", "copolymer", "adhesive"]):
        return "label_binder"
    if any(k in text for k in ["glass", "rg93", "ceramic"]):
        return "glass_phase"
    if any(k in text for k in ["alloy", "inconel", "ring", "pin", "cup", "plate", "hardgold", "ni(", "metal"]):
        return "metal_internal"
    if any(k in text for k in ["carbon black"]):
        return "form_specific"
    return "general"


def parse_hits(pages: List[str], rules_df: pd.DataFrame) -> pd.DataFrame:
    cas_set = set(rules_df["CAS No."].tolist())
    rules_by_cas = {row["CAS No."]: row.to_dict() for _, row in rules_df.iterrows()}

    current_l3 = ""
    current_l4 = ""
    current_material = ""
    hits: List[Dict[str, str]] = []

    for page_no, page in enumerate(pages, start=1):
        for raw in page.splitlines():
            line = normalize_text(raw)
            if not line:
                continue

            if line.startswith("3 "):
                current_l3 = line[2:]
            elif line.startswith("4 "):
                current_l4 = line[2:]
                current_material = current_l4
            elif line.startswith("5 "):
                current_material = line[2:]

            found = CAS_RE.findall(line)
            for cas in found:
                if cas not in cas_set:
                    continue
                rule = rules_by_cas[cas]
                substance_imds = re.sub(r"^\d+\s+", "", line.split(cas)[0]).strip()
                context = infer_context(current_l3, current_material, line)

                final_decision = rule.get("Default Final Decision", "Review required")
                external_position = rule.get("Default Recommended External Position", "")
                support = rule.get("Default Support Note", "")
                status_rule = rule.get("Prop 65 status / rule point", "")

                # Context-aware overrides
                if rule.get("Category") == "booster_propellant":
                    final_decision = "Additional support recommended before final external declaration"
                elif rule.get("Category") == "form_specific" and context != "form_specific":
                    final_decision = "No warning recommended"
                elif rule.get("Category") in {"metal_internal", "glass_phase", "label_binder"}:
                    final_decision = "No warning recommended"

                hits.append({
                    "No.": len(hits) + 1,
                    "Prop 65 chemical / evaluation item": rule.get("Prop 65 chemical / evaluation item", ""),
                    "CAS No.": cas,
                    "Exact IMDS material name": substance_imds or rule.get("Prop 65 chemical / evaluation item", ""),
                    "Component / location in IMDS": current_l3 or current_l4,
                    "IMDS evidence summary": line,
                    "Prop 65 status / rule point": status_rule,
                    "Final Decision": final_decision,
                    "Recommended support document / note": support,
                    "Recommended External Position": external_position,
                    "Rule Category": rule.get("Category", ""),
                    "Context Inference": context,
                    "Page": page_no,
                    "Raw Line": line,
                })

    return pd.DataFrame(hits)


def summarize_hits(hits_df: pd.DataFrame, meta: Dict[str, str]) -> pd.DataFrame:
    if hits_df.empty:
        return pd.DataFrame([{
            "Overall Result": "No matched rules",
            "Key Message": "No CAS hits matched the current Prop 65 rule set in the uploaded report.",
            "Recommended External Position": "",
            "Items Requiring Extra Support": ""
        }])

    has_followup = (hits_df["Final Decision"] == "Additional support recommended before final external declaration").any()
    overall = "Conditionally No Warning / 1 item for added support" if has_followup else "No warning recommended"

    followup_items = hits_df.loc[
        hits_df["Final Decision"] == "Additional support recommended before final external declaration",
        ["Prop 65 chemical / evaluation item", "Recommended support document / note"]
    ].drop_duplicates()

    items_text = " | ".join(
        f"{r['Prop 65 chemical / evaluation item']}: {r['Recommended support document / note']}"
        for _, r in followup_items.iterrows()
    )

    key_message = (
        "The report contains Proposition 65 listed chemicals or exact CAS matches. "
        "Most identified items appear as internal alloys, internal glass/ceramic phases, label/ink binders, or bound carbon black. "
        "For airbag applications, modules are ordinarily installed behind vehicle interior trim; even where a DAB is externally visible, the exposed part is the cover only."
    )

    recommended = (
        "Most matched items can be positioned as enclosed or bound forms based on the current report structure. "
        "Any booster/pyrotechnic lithium-carbonate item should be addressed separately with supporting evidence prior to final customer declaration."
    )

    return pd.DataFrame([{
        "Overall Result": overall,
        "Key Message": key_message,
        "Recommended External Position": recommended,
        "Items Requiring Extra Support": items_text
    }])


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = BORDER
    ws.row_dimensions[row].height = 24


def style_body(ws, start_row=2, row_height=36):
    for r in range(start_row, ws.max_row + 1):
        ws.row_dimensions[r].height = row_height
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER


def set_widths(ws, widths: Dict[str, int]):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_workbook(meta: Dict[str, str], summary_df: pd.DataFrame, hits_df: pd.DataFrame, refs_df: pd.DataFrame) -> bytes:
    wb = Workbook()

    # Cover
    ws = wb.active
    ws.title = "0_Cover"
    ws.merge_cells("A1:B1")
    ws["A1"] = "Prop 65 IMDS Assessment Report"
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        ("Document", meta.get("document", "")),
        ("Supplier", meta.get("supplier", "")),
        ("Part No.", meta.get("part_no", "")),
        ("Description", meta.get("description", "")),
        ("Assessment Conclusion", summary_df.iloc[0]["Overall Result"]),
        ("Core conclusion", summary_df.iloc[0]["Key Message"]),
        ("Follow-up item", summary_df.iloc[0]["Items Requiring Extra Support"]),
    ]
    rr = 2
    for k, v in rows:
        ws[f"A{rr}"] = k
        ws[f"B{rr}"] = v
        ws[f"A{rr}"].fill = SUB_FILL
        ws[f"A{rr}"].font = BOLD
        ws[f"A{rr}"].border = BORDER
        ws[f"B{rr}"].border = BORDER
        ws[f"A{rr}"].alignment = Alignment(vertical="top")
        ws[f"B{rr}"].alignment = Alignment(vertical="top", wrap_text=True)
        rr += 1
    set_widths(ws, {"A": 22, "B": 110})

    # Executive Summary
    ws = wb.create_sheet("1_Executive_Summary")
    ws.append(list(summary_df.columns))
    for _, row in summary_df.iterrows():
        ws.append(list(row.values))
    ws.append([])
    ws.append(["Document reviewed", f"{meta.get('description', '')} / Part No. {meta.get('part_no', '')}"])
    ws.append(["Assessment basis", "IMDS component/material content + Prop 65 applicability logic from the rule engine"])
    ws.append(["Note", "This workbook keeps exact IMDS material names wherever available."])
    style_header(ws)
    style_body(ws, row_height=42)
    set_widths(ws, {"A": 30, "B": 110, "C": 90, "D": 78})
    add_filter(ws)

    # Final Assessment
    ws = wb.create_sheet("2_Prop65_Final_Assessment")
    final_cols = [
        "No.", "Prop 65 chemical / evaluation item", "CAS No.", "Exact IMDS material name",
        "Component / location in IMDS", "IMDS evidence summary", "Prop 65 status / rule point",
        "Final Decision", "Recommended support document / note"
    ]
    ws.append(final_cols)
    final_df = hits_df[final_cols].copy()
    for _, row in final_df.iterrows():
        ws.append(list(row.values))
    style_header(ws)
    style_body(ws, row_height=54)
    set_widths(ws, {"A": 6, "B": 28, "C": 16, "D": 28, "E": 34, "F": 85, "G": 70, "H": 30, "I": 60})
    for r in range(2, ws.max_row + 1):
        txt = str(ws[f"H{r}"].value or "")
        if "No warning" in txt:
            ws[f"H{r}"].fill = GREEN_FILL
        elif "Additional support" in txt or "Review" in txt:
            ws[f"H{r}"].fill = YELLOW_FILL
    add_filter(ws)

    # Detailed Log
    ws = wb.create_sheet("3_Detailed_Assessment_Log")
    ws.append(list(hits_df.columns))
    for _, row in hits_df.iterrows():
        ws.append(list(row.values))
    style_header(ws)
    style_body(ws, row_height=40)
    set_widths(ws, {"A": 6, "B": 28, "C": 16, "D": 26, "E": 34, "F": 80, "G": 60, "H": 28, "I": 56, "J": 72, "K": 20, "L": 18, "M": 8, "N": 72})
    add_filter(ws)

    # CAS Extract
    ws = wb.create_sheet("4_IMDS_CAS_Extract")
    extract_cols = ["Page", "Component / location in IMDS", "Exact IMDS material name", "CAS No.", "Raw Line"]
    ws.append(extract_cols)
    for _, row in hits_df[extract_cols].iterrows():
        ws.append(list(row.values))
    style_header(ws)
    style_body(ws, row_height=34)
    set_widths(ws, {"A": 8, "B": 36, "C": 34, "D": 16, "E": 92})
    add_filter(ws)

    # Decision Logic
    ws = wb.create_sheet("5_Decision_Logic")
    ws.append(["English", "한국어"])
    logic_rows = [
        ("1. Exact CAS match alone does not automatically require a Proposition 65 warning.", "1. CAS 일치만으로 Prop 65 경고가 자동으로 필요한 것은 아닙니다."),
        ("2. The key question is whether the listed chemical can create reasonably foreseeable exposure in the listed form.", "2. 핵심 판단 기준은 해당 물질이 등재된 형태로 합리적으로 예측 가능한 노출을 일으킬 수 있는지 여부입니다."),
        ("3. Internal alloy, plating, glass-phase, and bound label/ink forms are generally handled as exposure-based items rather than automatic warning triggers.", "3. 내부 합금, 도금, 유리상, 결합된 라벨/잉크 형태는 자동 경고 대상이 아니라 노출 기반 항목으로 다룹니다."),
        ("4. Carbon black is form-specific and focuses on airborne, unbound respirable particles.", "4. Carbon black은 형태 의존적이며 공기 중 비결합 호흡성 입자 여부가 핵심입니다."),
        ("5. Lithium-carbonate in booster or pyrotechnic context should be treated as the main follow-up item, especially for post-deployment condition.", "5. booster 또는 pyrotechnic 맥락의 Lithium-carbonate는 특히 전개 후 상태를 중심으로 주요 후속 검토 항목으로 다뤄야 합니다."),
    ]
    for row in logic_rows:
        ws.append(list(row))
    ws.append([])
    ws.append(["Official source links used", "URL"])
    for _, row in refs_df.iterrows():
        ws.append([row["Title"], row["URL"]])
    style_header(ws)
    style_body(ws, row_height=38)
    ws["A8"].fill = SUB_FILL
    ws["A8"].font = BOLD
    ws["B8"].fill = SUB_FILL
    ws["B8"].font = BOLD
    set_widths(ws, {"A": 100, "B": 100})
    add_filter(ws)

    # Read Me
    ws = wb.create_sheet("6_Read_Me")
    ws.append(["English", "한국어"])
    rows = [
        ("1. Upload any MDS / IMDS PDF that contains extractable text.", "1. 텍스트 추출이 가능한 MDS / IMDS PDF를 업로드합니다."),
        ("2. Upload the rules workbook and adjust it as needed before analysis.", "2. 규칙 워크북을 업로드하고 필요 시 분석 전에 수정합니다."),
        ("3. Review 2_Prop65_Final_Assessment first for the customer-facing table.", "3. 고객 대응용 표는 먼저 2_Prop65_Final_Assessment 시트에서 확인합니다."),
        ("4. Use 3_Detailed_Assessment_Log and 4_IMDS_CAS_Extract for traceability.", "4. 추적성은 3_Detailed_Assessment_Log와 4_IMDS_CAS_Extract 시트에서 확인합니다."),
    ]
    for row in rows:
        ws.append(list(row))
    style_header(ws)
    style_body(ws, row_height=36)
    set_widths(ws, {"A": 88, "B": 88})
    add_filter(ws)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


st.title("Rule-Based Prop 65 IMDS Report Generator")

pdf_file = st.file_uploader("1) Upload MDS / IMDS PDF", type=["pdf"])
rule_file = st.file_uploader("2) Upload rule workbook (.xlsx)", type=["xlsx"])

st.markdown("### What this version improves")
st.write(
    "- Chemical rules are separated into an external Excel workbook.\n"
    "- You can add or revise chemicals without changing the Python code.\n"
    "- The report format stays consistent across similar MDS / IMDS PDFs.\n"
    "- Lithium-carbonate and other special logic can be maintained in the rules file."
)

if pdf_file and rule_file:
    with st.spinner("Parsing report and applying rules..."):
        pages = extract_pages(pdf_file.read())
        meta = find_metadata(pages)
        rules_df, refs_df = load_rules(rule_file.read())
        hits_df = parse_hits(pages, rules_df)
        summary_df = summarize_hits(hits_df, meta)
        xlsx_bytes = build_workbook(meta, summary_df, hits_df, refs_df)

    st.success("Report generated.")

    st.subheader("Preview - Executive Summary")
    st.dataframe(summary_df, use_container_width=True)

    st.subheader("Preview - Matched Chemicals")
    st.dataframe(
        hits_df[[
            "Prop 65 chemical / evaluation item",
            "CAS No.",
            "Component / location in IMDS",
            "Final Decision"
        ]].head(20),
        use_container_width=True,
    )

    st.download_button(
        "Download Excel Report",
        data=xlsx_bytes,
        file_name="Prop65_IMDS_Rule_Based_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("Upload both the PDF report and the rule workbook to begin.")
